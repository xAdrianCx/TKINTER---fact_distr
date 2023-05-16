from tkinter import (Tk, ttk, messagebox, StringVar, Label, Entry, Button, filedialog)
import tkinter.font
import customtkinter as ctk
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import pandas as pd
from configparser import ConfigParser
from threading import Thread

def get_excel():
	"""
	Gets the excel database needed to generate other files.
	:return:
	"""
	global input_file_path_var
	input_file_path_entry.delete(0, "end")
	excel_file = filedialog.askopenfilename(title="Select a file...", filetypes=[("Excel files", ".xlsx", ".xls")])
	input_file_path_var.set(excel_file)



def export_to():
	"""
	This function sets the folder where we want to generate the files.
	:return: Nothing
	"""
	global export_file_path_var
	export_file_path_entry.delete(0, "end")
	export_folder = filedialog.askdirectory(title="Choose directory...")
	export_file_path_var.set(export_folder)


def generate():
	"""
	This function creates a new .xlsx file for every supplier containing data to create an invoice.
	"""
	curr_sup_label.configure(text=f"Am inceput procesul...")
	year = year_var.get()
	month = month_var.get()
	year = year_var.get()

	# Get main app path.
	main_path = os.getcwd()
	# Get config gile path.
	cfg_path = os.path.join(main_path, "config\\cfg.ini")
	# Instantiate Configparser.
	config = ConfigParser()
	# Read config file.
	config.read(cfg_path)
	# Save config data.
	tariff_c1 = config.get("TARIFF", "tariff_c1")
	tariff_c2 = config.get("TARIFF", "tariff_c2")
	tariff_c3 = config.get("TARIFF", "tariff_c3")
	sup_ctr_no = dict(config.items("SUPPLIERS"))
	user = config.get("USERS", "user")

	# Set the path where the xlsx file is located.
	excel_path = input_file_path_var.get()
	# Get the folder where we save generated files.
	save_path = export_file_path_var.get()
	if excel_path == "":
		messagebox.showwarning("Warning!", "Trebuie sa alegeti un fisier de import excel(cu extensia .xlsx).")
	elif save_path == "":
		messagebox.showwarning("Warning!", "Alegeti locatia unde doriti sa salvati fisierele generate.")
	else:	
		# Get a data frame from excel file.
		df = pd.read_excel(excel_path, sheet_name="Sheet1", index_col=None, na_values=['NA'], usecols="A, E, L")
		df["Utilizatorul sistemului de distributie"] = df["Utilizatorul sistemului de distributie"].apply(str.upper)
		
		# Change current working directory to the folder where excel file is located.
		os.chdir(f"{save_path}")
		# Pivot the data frame.
		pivot = df.pivot_table(index=['Utilizatorul sistemului de distributie', 'Tipul categoriei'],
							   values=['MWh (cu 6 zecimale)'], aggfunc='sum')
		# Set a list to store quantities.
		total_mwh = []
		# Set a list to store suppliers and category
		sup_categ = []
		# Set a list to store only suppliers.
		sup_no_duplicates = []
		# Set a list to store suplier name, category and quantity(after pivoting).
		all_data = []

		# Save all quantities to total_mwh list.
		for i in pivot.values.tolist():
			total_mwh.append(i)

		# Save suplier name and category as a list to sup_categ list.
		for i in pivot.index:
			sup_categ.append(list(i))

		# Save suplier name, category to all_data list.
		for i in range(len(sup_categ)):
			all_data.append(sup_categ[i] + total_mwh[i])

		# Get all supplier names withut duplicates.
		for i in all_data:
			if i[0] not in sup_no_duplicates:
				sup_no_duplicates.append(i[0])

		# Create all needed excel files. Creates only files with name that we have data for.
		for i in all_data:
			wb = Workbook()
			ws = wb.active
			ws.title = "Sheet1"
			# FONTS.
			ft_black = Font(bold=True)
			ft_white = Font(bold=True, color="FFFFFF")
			# Write to cells.
			ws["D1"] = "Anexa 1"
			ws["A3"] = i[0]
			ws["C3"] = "Contract:"
			ws["C3"].alignment = Alignment(horizontal='right')
			for key, value in sup_ctr_no.items():
				if key.lower() == i[0].lower():
					ws["D3"] = value
					break
			ws["D3"].alignment = Alignment(horizontal='left')
			ws["A5"] = f"Facturare servicii de distributie pentru luna {month} {year}"
			ws["A5"].fill = PatternFill("solid", "0066CC")
			ws["A5"].font = ft_white
			ws["A5"].border = Border(top=Side(border_style='thick', color='FF000000'),
									 left=Side(border_style='thick', color='FF000000'),
									 bottom=Side(border_style='thick', color='FF000000'),
									 right=Side(border_style='thick', color='FF000000'))
			ws["A5"].alignment = Alignment(horizontal='center')
			ws.merge_cells("A5:D5")

			ws["A6"] = "Categorie"
			ws["A6"].font = ft_black
			ws["A6"].border = Border(top=Side(border_style='thick', color='FF000000'),
									 left=Side(border_style='thick', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["A6"].alignment = Alignment(horizontal='center')

			ws["B6"] = "Cantitate in MWh"
			ws["B6"].font = ft_black
			ws["B6"].border = Border(top=Side(border_style='thick', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["B6"].alignment = Alignment(horizontal='center')

			ws["C6"] = "Tarif de distributie"
			ws["C6"].font = ft_black
			ws["C6"].border = Border(top=Side(border_style='thick', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["C6"].alignment = Alignment(horizontal='center')

			ws["D6"] = "Valoare fara TVA"
			ws["D6"].font = ft_black
			ws["D6"].border = Border(top=Side(border_style='thick', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thick', color='FF000000'))
			ws["D6"].alignment = Alignment(horizontal='center')

			ws["A7"] = "C1"
			ws["A7"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thick', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["A7"].alignment = Alignment(horizontal='center')

			ws["B7"] = 0.000000
			ws["B7"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["B7"].alignment = Alignment(horizontal='center')

			ws["C7"] = tariff_c1
			ws["C7"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["C7"].alignment = Alignment(horizontal='center')

			ws["D7"] = "=B7*C7"
			ws["D7"].number_format = "0.00"
			ws["D7"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thick', color='FF000000'))
			ws["D7"].alignment = Alignment(horizontal='center')

			ws["A8"] = "C2"
			ws["A8"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thick', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["A8"].alignment = Alignment(horizontal='center')

			ws["B8"] = 0.000000
			ws["B8"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["B8"].alignment = Alignment(horizontal='center')

			ws["C8"] = tariff_c2
			ws["C8"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["C8"].alignment = Alignment(horizontal='center')

			ws["D8"] = "=B8*C8"
			ws["D8"].number_format = "0.00"
			ws["D8"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thin', color='FF000000'),
									 right=Side(border_style='thick', color='FF000000'))
			ws["D8"].alignment = Alignment(horizontal='center')

			ws["A9"] = "C3"
			ws["A9"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thick', color='FF000000'),
									 bottom=Side(border_style='thick', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["A9"].alignment = Alignment(horizontal='center')

			ws["B9"] = 0.000000
			ws["B9"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thick', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["B9"].alignment = Alignment(horizontal='center')

			ws["C9"] = tariff_c3
			ws["C9"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thick', color='FF000000'),
									 right=Side(border_style='thin', color='FF000000'))
			ws["C9"].alignment = Alignment(horizontal='center')

			ws["D9"] = "=B9*C9"
			ws["D9"].number_format = "0.00"
			ws["D9"].border = Border(top=Side(border_style='thin', color='FF000000'),
									 left=Side(border_style='thin', color='FF000000'),
									 bottom=Side(border_style='thick', color='FF000000'),
									 right=Side(border_style='thick', color='FF000000'))
			ws["D9"].alignment = Alignment(horizontal='center')

			ws["A10"] = "Total"
			ws["A10"].font = ft_black
			ws["A10"].alignment = Alignment(horizontal='center')
			ws["B10"] = "=SUM(B7:B9)"
			ws["B10"].number_format = "0.000000"
			ws["B10"].font = ft_black
			ws["B10"].alignment = Alignment(horizontal='center')
			ws["D10"] = "=SUM(D7:D9)"
			ws["D10"].number_format = "0.00"
			ws["D10"].font = ft_black
			ws["D10"].alignment = Alignment(horizontal='center')
			ws["A17"] = user
			ws["A17"].alignment = Alignment(horizontal='center')
			# Changing columns width.
			for col in ws.columns:
				column = col[0].column_letter
				adjusted_width = 21
				ws.column_dimensions[column].width = adjusted_width

			# Save and exit.
			wb.save(f"{i[0]} - {month} {year}.xlsx")
			curr_sup_label.configure(text=f"Generare fisier: {i[0]}")
			curr_sup_label.configure(text=f"Generare fisier: {i[0]}")
			

		# Open each file in directory and modify its content(if needed).
		for i in all_data:
			wb = load_workbook(f"{i[0]} - {month} {year}.xlsx")
			ws = wb.active
			if ws["A7"].value in i:
				ws["B7"] = float(f"{i[2]:.6f}")
				wb.save(f"{i[0]} - {month} {year}.xlsx")
			elif ws["A8"].value in i:
				ws["B8"] = float(f"{i[2]:.6f}")
				wb.save(f"{i[0]} - {month} {year}.xlsx")
			elif ws["A9"].value in i:
				ws["B9"] = float(f"{i[2]:.6f}")
				wb.save(f"{i[0]} - {month} {year}.xlsx")
		# Print a message at the end.
		curr_sup_label.configure(text="Proces terminat cu succes.")
		messagebox.showinfo("Info!", f"Au fost generate cu succes {len(sup_no_duplicates)} fisiere.")


def start_generate():
	"""
	This function starts the proces in a thread so we can do other stuff while processing files.
	"""
	Thread(target=generate, daemon=True).start()


# Set the main window.
root = ctk.CTk()
# Set main windows title.
root.title("Generator Facturi Distributie")
# Set main window width.
width = 520
# Set main window height
height = 380
# Set min size.
root.minsize(width, height)
# Set max size.
root.maxsize(width, height)

# Declare needed variables.
year_var = StringVar()
month_var = StringVar()
input_file_path_var = StringVar()
export_file_path_var = StringVar()

# Set a title at the top of the window.
title_font = ctk.CTkFont(family="Times New Roman", size=25, weight="bold", underline=True)
title = ctk.CTkLabel(root, text="GENERARE FACTURI DISTRIBUTIE", font=title_font)
title.grid(row=0, columnspan=3, padx=30, pady=30)
title.grid_rowconfigure(1, weight=1)
title.grid_columnconfigure(1, weight=1)

# Create year label and grid it to screen.
year_label = ctk.CTkLabel(root, text="Selectati anul:", font=("Times New Roman", 15))
year_label.grid(row=1, column=0, pady=10, padx=10, sticky="W")

# Create year combobox.
year_combo = ctk.CTkComboBox(root, 
				values=["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "2028", "2029", "2030",
								"2031", "2032", "2033", "2034", "2035", "2036", "2037", "2038", "2039", "2040",
								"2041", "2042", "2043", "2044", "2045", "2046", "2047", "2048", "2049", "2050"], 
				variable=year_var,
				state="readonly")
year_combo.grid(row=1, column=1, pady=10, padx=10)

# Create month label and grid it to screen.
month_label = ctk.CTkLabel(root, text="Selectati luna:", font=("Times New Roman", 15))
month_label.grid(row=2, column=0, pady=10, padx=10, sticky="W")
month_combo = ctk.CTkComboBox(root, 
				values=["ianuarie", "februarie", "martie", "aprilie", "mai", "iunie", "iulie", "august", "septembrie",
						 "octombrie", "noiembrie", "decembrie"], 
				variable=month_var, 
				state="readonly")
month_combo.grid(row=2, column=1, pady=10, padx=10)
# Set current year as default value.
for i in range(len(year_combo._values)):
	if datetime.today().year == int(year_combo._values[i]) and datetime.today().month == 1:
		year_combo.set(year_combo._values[i - 1])
	else:
		if datetime.today().year == int(year_combo._values[i]):
			year_combo.set(year_combo._values[i])

# Set current month to one month early than current month.
for i in range(0, len(month_combo._values)):
	if datetime.today().month == 1:
		month_combo.current(11)
	else:	
		month_combo.set(month_combo._values[(datetime.today().month) - 2])

# Create input file label, entrybox and askforfilepath button.
input_file_path_label = ctk.CTkLabel(root, text="Selectati fisierul de import:", font=("Times New Roman", 15))
input_file_path_label.grid(row=3, column=0, pady=10, padx=10, sticky="W")
input_file_path_entry = ctk.CTkEntry(root, textvariable=input_file_path_var, state="disabled")
input_file_path_entry.grid(row=3, column=1, pady=10, padx=10, sticky="W")
input_file_path_button = ctk.CTkButton(root, text="Alegeti fisierul...", font=("Times New Roman", 12),
									   text_color="white", command=get_excel)
input_file_path_button.grid(row=3, column=2, pady=10, padx=10, sticky="W")

# Create export file label, entry and askfordirectory button.
export_file_path_label = ctk.CTkLabel(root, text="Selectati directorul de export:", font=("Times New Roman", 15))
export_file_path_label.grid(row=4, column=0, pady=10, padx=10, sticky="W")
export_file_path_entry = ctk.CTkEntry(root, textvariable=export_file_path_var, state="disabled")
export_file_path_entry.grid(row=4, column=1, pady=10, padx=10, sticky="W")
export_file_path_button = ctk.CTkButton(root, text="Alegeti directorul...", font=("Times New Roman", 12),
										text_color="white", command=export_to)
export_file_path_button.grid(row=4, column=2, pady=10, padx=10, sticky="W")

# Crete a label that shows the current progress.
curr_sup_label = ctk.CTkLabel(root, text="", font=("Times New Roman bold", 15))
curr_sup_label.grid(row=6, columnspan=3, padx=10, pady=10, sticky="EW")

# Create start button.
start_button = ctk.CTkButton(root, text="Generare Fisiere", font=("Times New Roman", 20),
							 text_color="white", command=start_generate)
start_button.grid(row=5, columnspan=3, pady=10, padx=10)

# Run app.
root.mainloop()
