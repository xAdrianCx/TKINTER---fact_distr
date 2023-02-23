import os
import pandas as pd
from openpyxl import load_workbook, Workbook 
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, Color, colors
from configparser import ConfigParser 




month = "ianuarie"
year = 2023

tarif_c1 = 38.03
tarif_c2 = 34.52
tarif_c3 = 29.19

# Get main app path.
main_path = os.getcwd()
# Get config gile path.
cfg_path = os.path.join(main_path, "config\\cfg.ini")
# Instantiate Configparser.
config = ConfigParser()
# Read config file.
config.read(cfg_path)
# Save config data.
tariff_c1 = config.get("TARIFF", "tariff_c1")
tariff_c2 = config.get("TARIFF", "tariff_c2")
tariff_c3 = config.get("TARIFF", "tariff_c3")
sup_ctr_no = dict(config.items("SUPPLIERS"))

# Set the path where the xlsx file is located.
excel_path = "C:\\Users\\User\\Desktop\\New folder\\input\\01. Anexa 3 Ianuarie 2023 - copy.xlsx"
# Get a data frame from excel file. 
df = pd.read_excel(excel_path, sheet_name="Sheet1", index_col=None, na_values=['NA'], usecols="A, E, L")
# Get the folder where we save generated files.
save_path = os.path.join(main_path, "output")
# Change current working directory to the folder where excel file is located.
os.chdir(f"{save_path}")
# Pivot the data frame.
pivot = df.pivot_table(index=['Utilizatorul sistemului de distributie', 'Tipul categoriei'], values=['MWh (cu 6 zecimale)'], aggfunc='sum')
# Set a list to store quantities.
total_mwh = []
# Set a list to store suppliers and category
sup_categ = []
# Set a list to store only suppliers.
sup_no_duplicates = []
# Set a list to store suplier name, category and quantity(after pivoting).
all_data = []

# Save allquantities to total_mwh list.
for i in pivot.values.tolist():
	total_mwh.append(i)

# Save suplier name and category as a list to sup_categ list.
for i in pivot.index:
	sup_categ.append(list(i))

# save suplier name, category and to all_data list.
for i in range(len(sup_categ)):
	all_data.append(sup_categ[i] + total_mwh[i])

# Create all needed excel files. Creates only files with name that we have data for.
for i in all_data:
	wb = Workbook()
	ws = wb.active
	ws.title = "Sheet1"
	# FONTS.
	ft_black = Font(bold=True)
	ft_white = Font(bold=True, color="FFFFFF")
	
	ws["D1"] = "Anexa 1"
	ws["A3"] = i[0]
	ws["C3"] = "Contract:"
	ws["C3"].alignment = Alignment(horizontal='right')
	for key, value in sup_ctr_no.items():
		if key.lower() == i[0].lower():
			ws["D3"] = value
			break
	ws["D3"].alignment = Alignment(horizontal='left')
	ws["A5"] = f"Facturare servicii de distributie pentru luna {month} {year}"
	ws["A5"].fill = PatternFill("solid", "0066CC")
	ws["A5"].font = ft_white
	ws["A5"].border = Border(top=Side(border_style='thick', color='FF000000'),
                            left=Side(border_style='thick', color='FF000000'),
                            bottom=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='thick', color='FF000000'))
	ws["A5"].alignment = Alignment(horizontal='center')
	ws.merge_cells("A5:D5")

	ws["A6"] = "Categorie"
	ws["A6"].font = ft_black
	ws["A6"].border = Border(top=Side(border_style='thick', color='FF000000'),
                            left=Side(border_style='thick', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["A6"].alignment = Alignment(horizontal='center')

	ws["B6"] = "Cantitate in MWh"
	ws["B6"].font = ft_black
	ws["B6"].border = Border(top=Side(border_style='thick', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["B6"].alignment = Alignment(horizontal='center')

	ws["C6"] = "Tarif de distributie"
	ws["C6"].font = ft_black
	ws["C6"].border = Border(top=Side(border_style='thick', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["C6"].alignment = Alignment(horizontal='center')

	ws["D6"] = "Valoare fara TVA"
	ws["D6"].font = ft_black
	ws["D6"].border = Border(top=Side(border_style='thick', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thick', color='FF000000'))
	ws["D6"].alignment = Alignment(horizontal='center')

	ws["A7"] = "C1"
	ws["A7"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thick', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["A7"].alignment = Alignment(horizontal='center')

	ws["B7"] = 0.000000
	ws["B7"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["B7"].alignment = Alignment(horizontal='center')

	ws["C7"] = tariff_c1	
	ws["C7"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["C7"].alignment = Alignment(horizontal='center')

	ws["D7"] = "=B7*C7"
	ws["D7"].number_format = "0.00"
	ws["D7"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thick', color='FF000000'))
	ws["D7"].alignment = Alignment(horizontal='center')

	ws["A8"] = "C2"
	ws["A8"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thick', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["A8"].alignment = Alignment(horizontal='center')

	ws["B8"] = 0.000000
	ws["B8"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["B8"].alignment = Alignment(horizontal='center')

	ws["C8"] = tariff_c2
	ws["C8"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["C8"].alignment = Alignment(horizontal='center')

	ws["D8"] = "=B8*C8"
	ws["D8"].number_format = "0.00"
	ws["D8"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'),
                            right=Side(border_style='thick', color='FF000000'))
	ws["D8"].alignment = Alignment(horizontal='center')

	ws["A9"] = "C3"
	ws["A9"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thick', color='FF000000'),
                            bottom=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["A9"].alignment = Alignment(horizontal='center')

	ws["B9"] = 0.000000
	ws["B9"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["B9"].alignment = Alignment(horizontal='center')

	ws["C9"] = tariff_c3	
	ws["C9"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='thin', color='FF000000'))
	ws["C9"].alignment = Alignment(horizontal='center')

	ws["D9"] = "=B9*C9"
	ws["D9"].number_format = "0.00"
	ws["D9"].border = Border(top=Side(border_style='thin', color='FF000000'),
                            left=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='thick', color='FF000000'))
	ws["D9"].alignment = Alignment(horizontal='center')

	ws["A10"] = "Total"
	ws["A10"].font = ft_black
	ws["A10"].alignment = Alignment(horizontal='center')
	ws["B10"] = "=SUM(B7:B9)"
	ws["B10"].number_format = "0.000000"
	ws["B10"].font = ft_black
	ws["B10"].alignment = Alignment(horizontal='center')
	ws["D10"] = "=SUM(D7:D9)"
	ws["D10"].number_format = "0.00"
	ws["D10"].font = ft_black
	ws["D10"].alignment = Alignment(horizontal='center')

	ws["A17"] = "Cristian Orz"
	ws["A17"].alignment = Alignment(horizontal='center')

	# Save and exit.
	wb.save(f"{i[0]} {month} {year}.xlsx")
		
# Open each file in directory and modify its content(if needed).
for i in all_data:
	wb = load_workbook(f"{i[0]} {month} {year}.xlsx")
	ws = wb.active
	if ws["A7"].value in i:
		ws["B7"] = float(f"{i[2]:.6f}")
		wb.save(f"{i[0]} {month} {year}.xlsx")
	elif ws["A8"].value in i:	
		ws["B8"] = float(f"{i[2]:.6f}")
		wb.save(f"{i[0]} {month} {year}.xlsx")
	elif ws["A9"].value in i:
		ws["B9"] = float(f"{i[2]:.6f}")
		wb.save(f"{i[0]} {month} {year}.xlsx")
